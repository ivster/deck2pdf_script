import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Open the Excel file and select the active worksheet
path = '/Users/path/to/excel_file'
wb = load_workbook(path)
ws = wb.active


# Using Selenium to access website

# Initialize Chrome webdriver with options
chromeOptions = Options()
chromeOptions.add_argument("--incognito")

chromeOptions.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
chrome_driver_binary = "/Users/Local/Bin/chromedriver"
service = Service(executable_path=chrome_driver_binary)
driver = webdriver.Chrome(service=service, options=chromeOptions)

# Load the Deck2PDF website
driver.get('https://deck2pdf.com')

# Loop through the specified column of the worksheet and extract the links
for cell in ws['D']:
    link = cell.value.strip()

    # Open a new tab for each link
    driver.execute_script("window.open('');")
    driver.switch_to.window(driver.window_handles[-1])
    driver.get('https://deck2pdf.com')

    # Enter the link into the text box on the Deck2PDF website and submit the form
    input_box = driver.find_element(By.ID, 'docsendURL')
    input_box.clear()
    input_box.send_keys(link)
    submit_button = driver.find_element("xpath", "//button[contains(.,'Convert')]")
    submit_button.click()

    # Wait for the conversion to complete and download the PDF 
    try:
        wait = WebDriverWait(driver, 120)
        download_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='mainForm']/div/div/div/div/a")))
        download_button.click()

        # Wait for the download to complete before moving on to the next link
        time.sleep(5)

    except:
        # If there's an error while downloading the PDF, click convert instead
        convert_button = driver.find_element(By.XPATH, "//button[contains(.,'Convert')]")
        convert_button.click()
        continue
    
    # Press the Convert another button to put in another link
    wait_again = WebDriverWait(driver, 7)
    convert_another = wait_again.until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Convert')]")))
    convert_another.click()
    
    # Switch back to the original tab
    driver.switch_to.window(driver.window_handles[0])
    
    # Close the Chrome webdriver
    driver.quit()

    # Close the Excel workbook
    wb.close()
